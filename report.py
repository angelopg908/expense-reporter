import os
import logging
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from analytics import generate_kpis, spend_by_category, spend_by_month, top_expenses

logger = logging.getLogger(__name__)


def export_excel(df, output_path="output/expense_report.xlsx"):
    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    kpis = generate_kpis(df)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")

    ws_summary.append(["Metric", "Value"])
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for key, value in kpis.items():
        ws_summary.append([key.replace("_", " ").title(), value])

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    # --- Data sheet ---
    ws_data = wb.create_sheet(title="All Expenses")
    ws_data.append(["Date", "Description", "Category", "Amount", "Source File"])

    for cell in ws_data[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        ws_data.append([
            str(row["date"])[:10],
            str(row.get("description", "")),
            str(row.get("category", "")),
            float(row["amount"]),
            str(row.get("source_file", ""))
        ])

    for col in ["A", "B", "C", "D", "E"]:
        ws_data.column_dimensions[col].width = 22

    wb.save(output_path)
    logger.info(f"Excel report saved to {output_path}")


def export_dashboard(df, output_path="output/dashboard.html"):
    kpis = generate_kpis(df)
    by_category = spend_by_category(df)
    by_month = spend_by_month(df)
    top = top_expenses(df)

    fig = make_subplots(
    rows=2, cols=2,
    subplot_titles=(
        "Spend by Category",
        "Spend by Month",
        "Top 5 Expenses",
        "KPI Summary"
    ),
    specs=[
        [{"type": "pie"}, {"type": "xy"}],
        [{"type": "xy"}, {"type": "table"}]
    ]
    )

    # Spend by category - pie chart
    fig.add_trace(go.Pie(
        labels=list(by_category.keys()),
        values=list(by_category.values()),
        name="By Category"
    ), row=1, col=1)

    # Spend by month - bar chart
    fig.add_trace(go.Bar(
        x=list(by_month.keys()),
        y=list(by_month.values()),
        name="By Month",
        marker_color="#4472C4"
    ), row=1, col=2)

    # Top expenses - horizontal bar
    fig.add_trace(go.Bar(
        x=top["amount"].tolist(),
        y=top["description"].tolist(),
        orientation="h",
        name="Top Expenses",
        marker_color="#ED7D31"
    ), row=2, col=1)

    # KPI table
    fig.add_trace(go.Table(
        header=dict(values=["Metric", "Value"],
                    fill_color="#4472C4",
                    font=dict(color="white")),
        cells=dict(
            values=[
                [k.replace("_", " ").title() for k in kpis.keys()],
                list(kpis.values())
            ]
        )
    ), row=2, col=2)

    fig.update_layout(
        title_text="Monthly Expense Report",
        height=800,
        showlegend=False
    )

    fig.write_html(output_path)
    logger.info(f"Dashboard saved to {output_path}")